#!/usr/bin/env python3
"""
03_repooling_report.py
Automated repooling calculation for BRB-Seq spike-in analysis.

Replaces the manual copy-paste into Evenness_Repooling_Template.xlsx.

Reads:
  - STAR Log.final.out files (raw read counts, mapping stats)
  - FeatureCounts summary files (assigned read counts)
  - samples.tsv (well locations)
  - config.yaml (target reads, thresholds)

Outputs:
  - repooling_report.tsv    : Full per-sample metrics table
  - epmotion_export.tsv     : Ready to import into epMotion (same format as "Export Me" sheet)
  - repooling_summary.txt   : Human-readable summary with flags
  - repooling_template_filled.xlsx : Excel template with data pre-filled (optional)

Repooling formulas (reverse-engineered from Evenness_Repooling_Template.xlsx):
  projected_raw_M    = raw_K * target_M / sum(raw_K_all_samples)
  predicted_assigned = assigned_K * target_M / sum(raw_K_all_samples)
  amount_uL          = (min_vol_uL * max_raw_K) / raw_K

  where min_vol_uL = 2.0 (minimum pipettable volume)
  and target_M is the total reads requested for the full run (in millions)

Usage (called by submit_pipeline.sh):
  python3 03_repooling_report.py --config config.yaml

Or manually:
  python3 03_repooling_report.py \
      --project-dir /path/to/project \
      --samples-file metadata/samples.tsv \
      --target-reads 500 \
      --output-dir /path/to/project/repooling \
      --output-excel
"""

import argparse
import os
import sys
import yaml
import glob
import re
from pathlib import Path
try:
    import openpyxl
    from openpyxl.styles import PatternFill, Font, Alignment
    EXCEL_AVAILABLE = True
except ImportError:
    EXCEL_AVAILABLE = False


# ─────────────────────────────────────────────
# Constants
# ─────────────────────────────────────────────

MIN_VOL_UL = 2.0           # minimum pipettable volume
MIN_ASSIGNED_M = 4.0       # flag samples below this predicted assigned count (millions)
MAX_FOLD_CHANGE_WARN = 4.0 # warn if max/min assigned fold change exceeds this

# epMotion tool assignment thresholds (from Settings sheet)
TOOL_THRESHOLDS = [
    (10.0,  "TS_10"),
    (50.0,  "TS_50"),
    (300.0, "TS_300"),
    (float("inf"), "TS_1000"),
]


# ─────────────────────────────────────────────
# Argument parsing
# ─────────────────────────────────────────────

def parse_args():
    parser = argparse.ArgumentParser(description="BRB-Seq automated repooling report")
    parser.add_argument("--config", default="config_spikein.yaml",
                        help="Path to config.yaml (default: config.yaml)")
    # Allow overrides from command line (for calling from shell without config)
    parser.add_argument("--project-dir",  help="Override project directory from config")
    parser.add_argument("--samples-file", help="Override samples.tsv path from config")
    parser.add_argument("--target-reads", type=float,
                        help="Target total reads for full run (millions). Overrides config.")
    parser.add_argument("--output-dir",   help="Override output directory (default: project_dir/repooling)")
    parser.add_argument("--auto-target",  action="store_true",
                        help="Auto-optimize target_M to meet the 4M assigned threshold for most samples")
    parser.add_argument("--output-excel", action="store_true",
                        help="Generate Excel template file (requires openpyxl)")
    return parser.parse_args()


# ─────────────────────────────────────────────
# Config loading
# ─────────────────────────────────────────────

def load_config(config_path):
    if not os.path.exists(config_path):
        print(f"WARNING: Config file not found: {config_path}")
        return {}
    with open(config_path) as f:
        return yaml.safe_load(f)


# ─────────────────────────────────────────────
# Read samples.tsv
# ─────────────────────────────────────────────

def read_samples(samples_file):
    """Returns list of dicts with keys: SampleName, Barcode, Well"""
    samples = []
    with open(samples_file) as f:
        header = None
        for line in f:
            line = line.rstrip("\n\r")
            if not line:
                continue
            fields = line.split("\t")
            if header is None:
                header = [h.strip() for h in fields]
                continue
            row = dict(zip(header, fields))
            samples.append(row)
    return samples


# ─────────────────────────────────────────────
# Parse STAR log files
# ─────────────────────────────────────────────

def parse_star_log_final(star_log_path):
    """
    Parse STAR Log.final.out written by STAR.
    Returns dict with keys compatible with downstream code.
    """
    metrics = {
        "raw_reads": None,          # we'll use STAR input reads as proxy
        "trimmed_reads": None,
        "reads_input_star": None,
        "uniquely_mapped": None,
    }

    if not os.path.exists(star_log_path):
        return metrics

    with open(star_log_path) as f:
        for line in f:
            if "|" not in line:
                continue
            key, _, val = line.partition("|")
            key = key.strip()
            val = val.strip()

            if key == "Number of input reads":
                metrics["raw_reads"] = _to_int(val)
                metrics["reads_input_star"] = _to_int(val)
            elif key == "Uniquely mapped reads number":
                metrics["uniquely_mapped"] = _to_int(val)

    return metrics


def _to_int(val):
    """Extract integer from string, return None if not parseable."""
    val = re.sub(r"[^0-9]", "", val)
    return int(val) if val else None


# ─────────────────────────────────────────────
# Parse featureCounts summary files
# ─────────────────────────────────────────────

def parse_featurecounts_summary(summary_path):
    """
    Parse featureCounts .summary file.
    Returns dict with 'assigned' and 'unassigned' counts.
    """
    result = {"assigned": None}

    if not os.path.exists(summary_path):
        return result

    with open(summary_path) as f:
        for line in f:
            parts = line.strip().split("\t")
            if len(parts) >= 2 and parts[0] == "Assigned":
                result["assigned"] = _to_int(parts[1])

    return result


# ─────────────────────────────────────────────
# Core repooling calculations
# ─────────────────────────────────────────────

def compute_repooling(sample_data, target_M):
    """
    Given per-sample data and a target read depth, compute repooling metrics.

    sample_data: list of dicts with at minimum 'raw_reads_K' and 'assigned_reads_K'
    target_M: total reads requested for full run (millions)

    Returns updated list of dicts with repooling columns added.
    """
    # Only include samples with valid data
    valid = [s for s in sample_data
             if s.get("raw_reads_K") and s["raw_reads_K"] > 0
             and s.get("assigned_reads_K") is not None]

    if not valid:
        print("ERROR: No samples with valid read counts found.")
        sys.exit(1)

    total_raw_K = sum(s["raw_reads_K"] for s in valid)
    max_raw_K   = max(s["raw_reads_K"] for s in valid)

    for s in valid:
        raw_K      = s["raw_reads_K"]
        assigned_K = s["assigned_reads_K"] if s["assigned_reads_K"] else 0

        s["projected_raw_M"]       = raw_K * target_M / total_raw_K
        s["predicted_assigned_M"]  = assigned_K * target_M / total_raw_K
        s["amount_uL"]             = (MIN_VOL_UL * max_raw_K) / raw_K if raw_K > 0 else None
        s["tool"]                  = assign_tool(s["amount_uL"]) if s["amount_uL"] else None
        s["flag_low_assigned"]     = s["predicted_assigned_M"] < MIN_ASSIGNED_M
        s["flag_low_volume"]       = s["amount_uL"] is not None and s["amount_uL"] < MIN_VOL_UL

    # Compute fold change across all valid samples with assigned reads
    assigned_vals = [s["assigned_reads_K"] for s in valid if s.get("assigned_reads_K", 0) > 0]
    fold_change = max(assigned_vals) / min(assigned_vals) if assigned_vals and min(assigned_vals) > 0 else None

    return valid, total_raw_K, max_raw_K, fold_change


def assign_tool(volume_uL):
    """Return epMotion tool name based on volume."""
    for threshold, tool in TOOL_THRESHOLDS:
        if volume_uL <= threshold:
            return tool
    return "TS_1000"


def auto_optimize_target(sample_data, start_M=100, max_M=2000, step=50):
    """
    Find the minimum target_M where >= 90% of samples have predicted_assigned >= 4M.
    Returns the optimized target_M.
    """
    valid = [s for s in sample_data
             if s.get("raw_reads_K") and s["raw_reads_K"] > 0
             and s.get("assigned_reads_K") is not None]

    if not valid:
        return start_M

    total_raw_K = sum(s["raw_reads_K"] for s in valid)
    n = len(valid)

    for target_M in range(start_M, max_M + step, step):
        n_ok = sum(
            1 for s in valid
            if (s.get("assigned_reads_K", 0) or 0) * target_M / total_raw_K >= MIN_ASSIGNED_M
        )
        if n_ok / n >= 0.90:
            return target_M

    print(f"WARNING: Could not achieve 90% samples above {MIN_ASSIGNED_M}M assigned "
          f"within {max_M}M total reads. Using {max_M}M.")
    return max_M


# ─────────────────────────────────────────────
# Output writers
# ─────────────────────────────────────────────

def write_report_tsv(sample_data, output_path, target_M):
    """Write full per-sample metrics table."""
    headers = [
        "SampleName", "Well", "Raw_Reads_K", "Trimmed_Reads_K",
        "Mapped_K", "Pct_Mapped", "Assigned_K",
        "Projected_Raw_M", "Predicted_Assigned_M",
        "Amount_uL", "Tool",
        "Flag_Low_Assigned", "Flag_Low_Volume"
    ]

    with open(output_path, "w") as f:
        f.write("\t".join(headers) + "\n")
        for s in sample_data:
            pct_mapped = ""
            if s.get("uniquely_mapped") and s.get("reads_input_star"):
                pct_mapped = f"{s['uniquely_mapped'] / s['reads_input_star']:.3f}"

            row = [
                s.get("SampleName", ""),
                s.get("Well", ""),
                f"{s['raw_reads_K']:.1f}"          if s.get("raw_reads_K")          else "",
                f"{s['trimmed_reads_K']:.1f}"       if s.get("trimmed_reads_K")      else "",
                f"{s['mapped_K']:.1f}"              if s.get("mapped_K")             else "",
                pct_mapped,
                f"{s['assigned_reads_K']:.1f}"      if s.get("assigned_reads_K")     else "",
                f"{s['projected_raw_M']:.3f}"       if s.get("projected_raw_M")      else "",
                f"{s['predicted_assigned_M']:.3f}"  if s.get("predicted_assigned_M") else "",
                f"{s['amount_uL']:.3f}"             if s.get("amount_uL")            else "",
                s.get("tool", ""),
                "FLAG" if s.get("flag_low_assigned") else "",
                "FLAG" if s.get("flag_low_volume")   else "",
            ]
            f.write("\t".join(str(v) for v in row) + "\n")


def write_epmotion_tsv(sample_data, output_path):
    """
    Write epMotion-compatible export file.
    Matches the format of the 'Export Me' sheet in the Excel template.
    Columns: Rack, Source, Rack, Destination, Volume, Tool
    """
    with open(output_path, "w") as f:
        f.write("Rack\tSource\tRack\tDestination\tVolume\tTool\n")
        for s in sample_data:
            if not s.get("Well") or not s.get("amount_uL"):
                continue
            # Skip samples with 0 or missing reads
            if not s.get("raw_reads_K") or s["raw_reads_K"] == 0:
                continue
            f.write(f"1\t{s['Well']}\t1\t1\t{s['amount_uL']:.6f}\t{s['tool']}\n")


def write_summary(sample_data, output_path, target_M, total_raw_K, fold_change):
    """Write human-readable summary with flags."""
    n_total = len(sample_data)
    n_low_assigned = sum(1 for s in sample_data if s.get("flag_low_assigned"))
    n_low_volume   = sum(1 for s in sample_data if s.get("flag_low_volume"))
    n_high_volume  = sum(1 for s in sample_data
                         if s.get("amount_uL") and s["amount_uL"] > 50)

    with open(output_path, "w") as f:
        f.write("=" * 60 + "\n")
        f.write("BRB-Seq Spike-In Repooling Report\n")
        f.write("=" * 60 + "\n\n")

        f.write(f"Total samples:              {n_total}\n")
        f.write(f"Total raw reads (all):      {total_raw_K / 1000:.1f}M\n")
        f.write(f"Target reads (full run):    {target_M}M\n")
        f.write(f"Max/min assigned fold change: {fold_change:.1f}x\n")
        f.write("\n")

        # Fold change warning
        if fold_change and fold_change > MAX_FOLD_CHANGE_WARN:
            f.write(f"⚠ WARNING: Fold change ({fold_change:.1f}x) exceeds {MAX_FOLD_CHANGE_WARN}x.\n")
            f.write("  Consider repooling to equalize sample representation.\n\n")
        else:
            f.write(f"✓ Fold change ({fold_change:.1f}x) is within acceptable range (<{MAX_FOLD_CHANGE_WARN}x).\n\n")

        # Predicted assigned flags
        if n_low_assigned > 0:
            f.write(f"⚠ {n_low_assigned} sample(s) have predicted assigned counts < {MIN_ASSIGNED_M}M:\n")
            for s in sample_data:
                if s.get("flag_low_assigned"):
                    f.write(f"   {s['SampleName']} ({s.get('Well','?')}): "
                            f"{s.get('predicted_assigned_M', 0):.2f}M predicted\n")
            f.write("\n")
        else:
            f.write(f"✓ All samples have predicted assigned counts >= {MIN_ASSIGNED_M}M.\n\n")

        # Volume flags
        if n_high_volume > 0:
            f.write(f"⚠ {n_high_volume} sample(s) have repooling volume > 50 uL (TS_300 or higher):\n")
            for s in sample_data:
                if s.get("amount_uL") and s["amount_uL"] > 50:
                    f.write(f"   {s['SampleName']} ({s.get('Well','?')}): "
                            f"{s.get('amount_uL', 0):.1f} uL — consider capping manually.\n")
            f.write("\n")

        f.write("-" * 60 + "\n")
        f.write("Next steps:\n")
        f.write("  1. Review repooling_report.tsv for full per-sample metrics.\n")
        f.write("  2. If fold change > 4x, discuss repooling with Brian.\n")
        f.write("  3. If any sample amounts look unreasonably high (e.g. >100 uL),\n")
        f.write("     consider manually capping those volumes or excluding the sample.\n")
        f.write("  4. Import epmotion_export.tsv into epMotion RT_Repool_Template.\n")
        f.write("=" * 60 + "\n")


def write_excel_template(sample_data, output_path, target_M, samples_tsv):
    """
    Write an Excel file matching the Evenness_Repooling_Template structure.
    Fills in MultiQC_Report, Mapping, and Export Me sheets.
    """
    if not EXCEL_AVAILABLE:
        print("ERROR: openpyxl not installed. Cannot generate Excel output.")
        print("Install with: pip install openpyxl --break-system-packages")
        return False

    wb = openpyxl.Workbook()
    
    # Remove default sheet
    wb.remove(wb.active)
    
    # ─────────────────────────────────────────────
    # Sheet 1: MultiQC_Report
    # ─────────────────────────────────────────────
    
    ws_mqc = wb.create_sheet("MultiQC_Report")
    
    # Instructions (rows 1-6)
    ws_mqc['B1'] = "Instructions For Use:"
    ws_mqc['B2'] = "This file has been auto-generated by the BRB-Seq pipeline repooling script."
    ws_mqc['B3'] = "All sample data and calculations are pre-filled."
    ws_mqc['B4'] = f"Current target read depth: {target_M}M total reads"
    ws_mqc['B5'] = "To adjust target depth, edit cell K106 and calculations will update."
    ws_mqc['B6'] = "Export the 'Export Me' sheet as CSV for epMotion."
    
    # Header row (row 7)
    headers = [
        "SampleNumber", "Sample Name", "Raw Reads (K)", "Trimmed (K)", 
        "Length (bp)", "% Dup", "Mapped (K)", "% Mapped", "Assigned (K)", 
        None, "Projected Raw Reads (Millions)", None, 
        "Predicted Assigned counts (Green = > 4 million)", None, 
        "Repooling Amount\n(uL)", "Well Number"
    ]
    
    for col_idx, header in enumerate(headers, start=1):
        if header:
            cell = ws_mqc.cell(row=7, column=col_idx, value=header)
            cell.font = Font(bold=True)
            cell.alignment = Alignment(wrap_text=True, horizontal='center')
    
    # Data rows (starting row 8)
    for i, s in enumerate(sample_data, start=1):
        row = 7 + i
        
        # Calculate percentage mapped
        pct_mapped = ""
        if s.get("uniquely_mapped") and s.get("reads_input_star"):
            pct_mapped = s["uniquely_mapped"] / s["reads_input_star"]
        
        ws_mqc.cell(row=row, column=1, value=i)  # SampleNumber
        ws_mqc.cell(row=row, column=2, value=s.get("SampleName", ""))
        ws_mqc.cell(row=row, column=3, value=s.get("raw_reads_K"))
        ws_mqc.cell(row=row, column=4, value=s.get("trimmed_reads_K"))
        ws_mqc.cell(row=row, column=5, value="150 bp")  # Assumed
        ws_mqc.cell(row=row, column=6, value="")  # % Dup - not calculated
        ws_mqc.cell(row=row, column=7, value=s.get("mapped_K"))
        ws_mqc.cell(row=row, column=8, value=pct_mapped)
        ws_mqc.cell(row=row, column=9, value=s.get("assigned_reads_K"))
        ws_mqc.cell(row=row, column=10, value=None)  # Empty column
        ws_mqc.cell(row=row, column=11, value=s.get("projected_raw_M"))
        ws_mqc.cell(row=row, column=12, value=None)  # Empty column
        ws_mqc.cell(row=row, column=13, value=s.get("predicted_assigned_M"))
        ws_mqc.cell(row=row, column=14, value=None)  # Empty column
        ws_mqc.cell(row=row, column=15, value=s.get("amount_uL"))
        ws_mqc.cell(row=row, column=16, value=s.get("Well", ""))
        
        # Conditional formatting for predicted assigned (green if >= 4M)
        pred_cell = ws_mqc.cell(row=row, column=13)
        if s.get("predicted_assigned_M") and s["predicted_assigned_M"] >= MIN_ASSIGNED_M:
            pred_cell.fill = PatternFill(start_color="C6EFCE", end_color="C6EFCE", fill_type="solid")
    
    # Add target reads input cell (K106 in original template)
    summary_row = 7 + len(sample_data) + 3
    ws_mqc.cell(row=summary_row, column=11, value=target_M)
    ws_mqc.cell(row=summary_row-1, column=11, value="Number Reads Requested (Millions):")
    
    # ─────────────────────────────────────────────
    # Sheet 2: Mapping
    # ─────────────────────────────────────────────
    
    ws_map = wb.create_sheet("Mapping")
    ws_map['A1'] = "Well Location"
    ws_map['B1'] = "Sample Name"
    ws_map.cell(1, 1).font = Font(bold=True)
    ws_map.cell(1, 2).font = Font(bold=True)
    
    for i, s in enumerate(sample_data, start=2):
        ws_map.cell(row=i, column=1, value=s.get("Well", ""))
        ws_map.cell(row=i, column=2, value=s.get("SampleName", ""))
    
    # ─────────────────────────────────────────────
    # Sheet 3: Export Me
    # ─────────────────────────────────────────────
    
    ws_export = wb.create_sheet("Export Me")
    export_headers = ["Rack", "Source", "Rack", "Destination", "Volume", "Tool", "Name"]
    for col_idx, header in enumerate(export_headers, start=1):
        cell = ws_export.cell(row=1, column=col_idx, value=header)
        cell.font = Font(bold=True)
    
    row_idx = 2
    for s in sample_data:
        if not s.get("Well") or not s.get("amount_uL"):
            continue
        if not s.get("raw_reads_K") or s["raw_reads_K"] == 0:
            continue
        
        ws_export.cell(row=row_idx, column=1, value=1)  # Rack
        ws_export.cell(row=row_idx, column=2, value=s["Well"])  # Source
        ws_export.cell(row=row_idx, column=3, value=1)  # Rack
        ws_export.cell(row=row_idx, column=4, value="1")  # Destination
        ws_export.cell(row=row_idx, column=5, value=s["amount_uL"])  # Volume
        ws_export.cell(row=row_idx, column=6, value=s.get("tool", ""))  # Tool
        ws_export.cell(row=row_idx, column=7, value=None)  # Name (empty)
        row_idx += 1
    
    # ─────────────────────────────────────────────
    # Sheet 4: Settings
    # ─────────────────────────────────────────────
    
    ws_settings = wb.create_sheet("Settings")
    ws_settings['A1'] = "Below Volume"
    ws_settings['B1'] = "Tool"
    ws_settings.cell(1, 1).font = Font(bold=True)
    ws_settings.cell(1, 2).font = Font(bold=True)
    
    settings_data = [
        (0, "TS_10"),
        (0.1, "TS_10"),
        (10.1, "TS_50"),
        (50.1, "TS_300"),
        (300.1, "TS_1000"),
    ]
    
    for row_idx, (volume, tool) in enumerate(settings_data, start=2):
        ws_settings.cell(row=row_idx, column=1, value=volume)
        ws_settings.cell(row=row_idx, column=2, value=tool)
    
    # Save workbook
    try:
        wb.save(output_path)
        return True
    except Exception as e:
        print(f"ERROR: Failed to save Excel file: {e}")
        return False


# ─────────────────────────────────────────────
# Main
# ─────────────────────────────────────────────

def main():
    args = parse_args()

    # Load config
    config = load_config(args.config)

    # Resolve paths
    project_dir  = args.project_dir  or config.get("project", {}).get("directory", ".")
    samples_file = args.samples_file or config.get("input",   {}).get("samples_file", "metadata/samples.tsv")
    target_M     = args.target_reads or config.get("repooling", {}).get("target_reads_per_sample", 500)
    output_dir   = args.output_dir   or os.path.join(project_dir, "repooling")

    os.makedirs(output_dir, exist_ok=True)

    print(f"Project directory: {project_dir}")
    print(f"Samples file:      {samples_file}")
    print(f"Target reads:      {target_M}M")
    print(f"Output directory:  {output_dir}")
    print()

    # Read sample list
    if not os.path.exists(samples_file):
        print(f"ERROR: Samples file not found: {samples_file}")
        sys.exit(1)

    samples = read_samples(samples_file)
    print(f"Found {len(samples)} samples in {samples_file}")

    # Collect per-sample metrics
    sample_data = []
    missing_logs = []
    missing_fc   = []

    for s in samples:
        name = s["SampleName"]
        well = s.get("Well", "")

        # STAR log file
        star_log_path = os.path.join(project_dir, "STAR", f"{name}_Log.final.out")
        log_metrics = parse_star_log_final(star_log_path)
        if log_metrics["raw_reads"] is None:
            missing_logs.append(name)

        # FeatureCounts summary
        fc_pattern = os.path.join(project_dir, "FeatureCounts", f"{name}_featureCounts.txt.summary")
        fc_matches = glob.glob(fc_pattern)
        if fc_matches:
            fc_metrics = parse_featurecounts_summary(fc_matches[0])
        else:
            fc_metrics = {"assigned": None}
            missing_fc.append(name)

        # Convert to thousands (K) to match Excel template
        raw_K      = log_metrics["raw_reads"]      / 1000 if log_metrics["raw_reads"]      else None
        trimmed_K  = log_metrics["trimmed_reads"]  / 1000 if log_metrics["trimmed_reads"]  else None
        mapped_K   = log_metrics["uniquely_mapped"]/ 1000 if log_metrics["uniquely_mapped"] else None
        assigned_K = fc_metrics["assigned"]        / 1000 if fc_metrics["assigned"]         else None

        sample_data.append({
            "SampleName":      name,
            "Well":            well,
            "raw_reads_K":     raw_K,
            "trimmed_reads_K": trimmed_K,
            "reads_input_star":log_metrics["reads_input_star"],
            "uniquely_mapped": log_metrics["uniquely_mapped"],
            "mapped_K":        mapped_K,
            "assigned_reads_K":assigned_K,
        })

    # Warn about missing files
    if missing_logs:
        print(f"WARNING: Log files not found for {len(missing_logs)} sample(s):")
        for name in missing_logs:
            print(f"  {name}")
    if missing_fc:
        print(f"WARNING: FeatureCounts summary not found for {len(missing_fc)} sample(s):")
        for name in missing_fc:
            print(f"  {name}")

    # Auto-optimize target if requested
    if args.auto_target:
        original_target = target_M
        target_M = auto_optimize_target(sample_data)
        print(f"Auto-optimized target: {original_target}M → {target_M}M")

    # Compute repooling metrics
    computed, total_raw_K, max_raw_K, fold_change = compute_repooling(sample_data, target_M)

    print(f"\nTotal raw reads (all samples): {total_raw_K / 1000:.1f}M")
    print(f"Max/min assigned fold change:  {fold_change:.1f}x")
    if fold_change and fold_change > MAX_FOLD_CHANGE_WARN:
        print(f"WARNING: Fold change exceeds {MAX_FOLD_CHANGE_WARN}x — consider repooling.")

    # Write outputs
    report_path   = os.path.join(output_dir, "repooling_report.tsv")
    epmotion_path = os.path.join(output_dir, "epmotion_export.tsv")
    summary_path  = os.path.join(output_dir, "repooling_summary.txt")

    write_report_tsv(computed, report_path, target_M)
    write_epmotion_tsv(computed, epmotion_path)
    write_summary(computed, summary_path, target_M, total_raw_K, fold_change)

    print(f"\nOutputs written to {output_dir}/:")
    print(f"  repooling_report.tsv   — full per-sample metrics table")
    print(f"  epmotion_export.tsv    — import into epMotion RT_Repool_Template")
    print(f"  repooling_summary.txt  — human-readable summary with flags")
    
    # Generate Excel file if requested
    if args.output_excel:
        excel_path = os.path.join(output_dir, "repooling_template_filled.xlsx")
        print(f"\nGenerating Excel template...")
        success = write_excel_template(computed, excel_path, target_M, samples_file)
        if success:
            print(f"  repooling_template_filled.xlsx — Excel template with data pre-filled")
        else:
            print(f"  WARNING: Excel generation failed")

    # Print summary to stdout too
    print()
    with open(summary_path) as f:
        print(f.read())


if __name__ == "__main__":
    main()
